from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment
import re

print("Process Starts")

# Load existing workbook if it exists
existing_filename = 'Docs/TC_Summary.xlsx'
try:
    existing_workbook = openpyxl.load_workbook(existing_filename)
except FileNotFoundError:
    existing_workbook = None

# Create an Excel workbook and define the filename
workbook = openpyxl.Workbook()
filename = 'Docs/TC_Summary.xlsx'

app_html = 'Docs/Test_Plan_HTML/allclusters.html'
main_html = 'Docs/Test_Plan_HTML/index.html'

# Create or load the "All_TC_Details" sheet
sheet_name = "All_TC_Details"
if existing_workbook and sheet_name in existing_workbook.sheetnames:
    sheet1 = existing_workbook[sheet_name]
else:
    sheet1 = workbook.active
    sheet1.title = sheet_name

    # Define column headers
    headers = ['S.No', 'Cluster Name', 'Test Case Name', 'Test Case ID', 'Test Plan']

    # Add headers to the first row and set the font to bold for the headings
    header_font = Font(name='Times New Roman', bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = sheet1.cell(row=1, column=col_num, value=header)
        cell.font = header_font

    # Set header row alignment to center
    for cell in sheet1[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Create or load the "TC_Changes" sheet
change_sheet_name = "TC_Changes"
if existing_workbook and change_sheet_name in existing_workbook.sheetnames:
    change_sheet = existing_workbook[change_sheet_name]
else:
    change_sheet = workbook.create_sheet(title=change_sheet_name)

    # Define column headers for TC_Changes sheet
    change_headers = ['Action', 'S.No', 'Cluster Name', 'Test Case Name', 'Test Case ID', 'Test Plan']

    # Add headers to the first row and set the font to bold for the headings
    header_font = Font(name='Times New Roman', bold=True)
    for col_num, header in enumerate(change_headers, 1):
        cell = change_sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font

    # Set header row alignment to center
    for cell in change_sheet[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Parse 'app' HTML
print("^" * 40)
print("Parsing 'app' HTML...")
with open(app_html, encoding='utf-8') as f1:
    soup1 = BeautifulSoup(f1, 'html.parser')
    h1_tags1 = soup1.find_all('h1', {'id': True})
    extract_tc_details(h1_tags1, 1, 1, sheet1)

# Calculate the next row_number after parsing the first HTML
row_number = sheet1.max_row + 1

# Parse 'main' HTML
print("^" * 40)
print("Parsing 'main' HTML...")
with open(main_html, encoding='utf-8') as f2:
    soup2 = BeautifulSoup(f2, 'html.parser')
    h1_tags2 = soup2.find_all('h1', {'id': True})
    extract_tc_details(h1_tags2, 0, row_number, sheet1)

# Set the font for the entire sheet to Times New Roman
for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
    for cell in row:
        cell.font = Font(name='Times New Roman')
        cell.alignment = Alignment(vertical='center')  # Center-align vertically

# Set alignment to center for columns A and E
for column_letter in ['A', 'E']:
    for cell in sheet1[column_letter]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Set column widths
column_widths = {'A': 5, 'B': 30, 'C': 100, 'D': 25, 'E': 15}
for column, width in column_widths.items():
    sheet1.column_dimensions[column].width = width

# Extracted test case details
extracted_data = []

# Compare with existing data
existing_data = []
for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
    row_data = [cell.value for cell in row]
    existing_data.append(row_data)

# Identify added and removed test cases
added_test_cases = [row for row in extracted_data if row not in existing_data]
removed_test_cases = [row for row in existing_data if row not in extracted_data]

# Write added test cases to "TC_Changes" sheet
for added_row in added_test_cases:
    change_sheet.append(['Added'] + added_row)

# Write removed test cases to "TC_Changes" sheet
for removed_row in removed_test_cases:
    change_sheet.append(['Removed'] + removed_row)

# Save the workbook
print("Saving Excel workbook...")
workbook.save(filename)

print("Process completed. Excel file saved as 'TC_Summary.xlsx'.")
